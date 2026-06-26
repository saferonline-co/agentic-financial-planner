"""Excel workbook in the user's `20yr financial plan.xlsx` layout.

Years as columns; rows grouped Assets / Liabilities / Net assets -> Inflows /
Outflows -> Net inflow (cash-flow) -> Net worth -> today's money. Plus a Summary
(scenario comparison) and a MonteCarlo sheet.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..analysis import ScenarioResult

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
GOOD = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")
TODAY_FILL = PatternFill("solid", fgColor="FFF2CC")
NUMFMT = "#,##0"
CENTER = Alignment(horizontal="center")


def _section(ws, row, label):
    c = ws.cell(row, 1, label)
    c.font = SUB
    c.fill = SUB_FILL
    return row + 1


def _line(ws, row, label, values, col0=2, fmt=NUMFMT, fill=None):
    ws.cell(row, 1, label)
    for i, v in enumerate(values):
        cell = ws.cell(row, col0 + i, v)
        cell.number_format = fmt
        if fill:
            cell.fill = fill
    return row + 1


def _ledger_sheet(wb, result: ScenarioResult):
    cfg, rows = result.cfg, result.rows
    name = f"ledger_{result.name}"[:31]
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 26
    years = [r.year for r in rows]

    # header row of years
    ws.cell(1, 1, f"{result.name} — annual ledger (AUD, nominal)").font = SUB
    r = 2
    ws.cell(r, 1, "Year").font = HDR
    ws.cell(r, 1).fill = HDR_FILL
    for i, y in enumerate(years):
        c = ws.cell(r, 2 + i, y)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = CENTER
    r += 1

    def col(attr, transform=lambda x: x):
        return [transform(getattr(row, attr)) for row in rows]

    def acct(name):
        return [row.accounts.get(name, 0) for row in rows]

    r = _section(ws, r, "ASSETS AT START")
    r = _line(ws, r, "Cash", acct("cash"))
    r = _line(ws, r, "ETF / investments", acct("etf"))
    r = _line(ws, r, "Super (Person 1)", acct("person1_super"))
    r = _line(ws, r, "Super (Person 2)", acct("person2_super"))
    r = _line(ws, r, "Foreign pension (Person 1)", acct("person1_foreign_pension"))
    r = _line(ws, r, "Foreign pension (Person 2)", acct("person2_foreign_pension"))
    r = _line(ws, r, "Tax-free (Person 1)", acct("person1_taxfree"))
    r = _line(ws, r, "House", col("house_value"))
    r = _section(ws, r, "LIABILITIES")
    r = _line(ws, r, "Mortgage", col("mortgage"))
    r = _section(ws, r, "NET ASSETS")
    r = _line(ws, r, "Net worth at end", col("net_worth_nominal"))

    r = _section(ws, r, "INFLOWS")
    r = _line(ws, r, "Take-home pay", col("take_home"))
    r = _line(ws, r, "Foreign state pension", col("pension_income"))
    r = _line(ws, r, "Inheritance", col("inheritance"))
    r = _section(ws, r, "OUTFLOWS")
    r = _line(ws, r, "Living expenses", col("living"))
    r = _line(ws, r, "Housing (rent/mortgage)", col("housing"))
    r = _line(ws, r, "School", col("school"))
    r = _line(ws, r, "Tax", col("tax_paid"))
    r = _line(ws, r, "Aged care", col("care"))
    r = _line(ws, r, "Retirement spend (need)", col("need"))
    r = _line(ws, r, "Portfolio withdrawal", col("withdrawal"))

    r = _section(ws, r, "CASH FLOW & NET WORTH")
    r = _line(ws, r, "Net inflow through year", col("net_cashflow"))
    r = _line(ws, r, "Net worth at end (nominal)", col("net_worth_nominal"))
    r = _line(ws, r, "Net worth (today's money)", col("net_worth_today"), fill=TODAY_FILL)
    ws.freeze_panes = "B3"


def _summary_sheet(wb, results: list[ScenarioResult]):
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 24
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 14
    heads = ["Scenario", "P(sustains)", "Capital preserved", "Funded ratio",
             "Terminal P10", "Terminal P50", "Terminal P90", "Det. terminal"]
    for j, h in enumerate(heads):
        c = ws.cell(1, j + 1, h)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = CENTER
    for i, res in enumerate(results):
        row = i + 2
        ws.cell(row, 1, res.name)
        c = ws.cell(row, 2, res.mc.p_sustains)
        c.number_format = "0%"
        c.fill = GOOD if res.mc.p_sustains >= 0.90 else (BAD if res.mc.p_sustains < 0.80 else TODAY_FILL)
        ws.cell(row, 3, res.mc.p_capital_preserved).number_format = "0%"
        ws.cell(row, 4, res.funded_ratio).number_format = "0.00"
        for k, v in enumerate([res.mc.terminal_p10, res.mc.terminal_p50,
                               res.mc.terminal_p90, res.terminal_today]):
            ws.cell(row, 5 + k, v).number_format = NUMFMT


def _montecarlo_sheet(wb, results: list[ScenarioResult]):
    ws = wb.create_sheet("MonteCarlo")
    ws.column_dimensions["A"].width = 24
    heads = ["Scenario", "Trials", "P(sustains)", "Capital preserved",
             "P10 (today)", "P50 (today)", "P90 (today)", "Median depletion yr"]
    for j, h in enumerate(heads):
        c = ws.cell(1, j + 1, h)
        c.font = HDR
        c.fill = HDR_FILL
    for i, res in enumerate(results):
        m = res.mc
        row = i + 2
        ws.cell(row, 1, res.name)
        ws.cell(row, 2, m.trials)
        ws.cell(row, 3, m.p_sustains).number_format = "0%"
        ws.cell(row, 4, m.p_capital_preserved).number_format = "0%"
        ws.cell(row, 5, m.terminal_p10).number_format = NUMFMT
        ws.cell(row, 6, m.terminal_p50).number_format = NUMFMT
        ws.cell(row, 7, m.terminal_p90).number_format = NUMFMT
        ws.cell(row, 8, m.median_depletion_year or "-")


def write_report(path: str, results: list[ScenarioResult]):
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    _summary_sheet(wb, results)
    _montecarlo_sheet(wb, results)
    for res in results:
        _ledger_sheet(wb, res)
    wb.save(path)
    return path
