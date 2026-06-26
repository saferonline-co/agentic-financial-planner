"""TDD: analysis bundling + markdown summary + Excel workbook."""
import openpyxl
import pytest
from finmodel import schema, analysis
from finmodel.report import markdown, excel


@pytest.fixture
def result():
    cfg = schema.load_config()
    return analysis.analyse(cfg, scenario_name="recommended", trials=200)


def test_analyse_bundles_everything(result):
    assert result.name == "recommended"
    assert result.rows[0].year == 2026
    assert 0 <= result.mc.p_sustains <= 1
    assert result.funded_ratio > 0
    assert result.terminal_today != 0


def test_markdown_summary_has_key_sections(result):
    md = markdown.summary(result)
    assert "# " in md                       # a title
    assert "sustain" in md.lower()
    assert "funded ratio" in md.lower()
    assert "net worth" in md.lower()
    assert "2072" in md                      # horizon referenced


def test_excel_workbook_has_expected_sheets(tmp_path, result):
    out = tmp_path / "recommended.xlsx"
    excel.write_report(str(out), [result])
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert any("recommended" in s for s in wb.sheetnames)
    assert "MonteCarlo" in wb.sheetnames


def test_excel_ledger_has_cashflow_and_networth_rows(tmp_path, result):
    out = tmp_path / "recommended.xlsx"
    excel.write_report(str(out), [result])
    wb = openpyxl.load_workbook(out)
    ws = next(wb[s] for s in wb.sheetnames if "recommended" in s)
    labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(l and "Net inflow" in str(l) for l in labels)
    assert any(l and "Net worth at end" in str(l) for l in labels)
    assert any(l and "today" in str(l).lower() for l in labels)
